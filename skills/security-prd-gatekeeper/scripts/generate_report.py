"""
安全设计评审报告生成脚本

基于团队统一的《信息系统安全设计Checklist》模板，把评审结果写回 Excel，
并额外生成一个“评审摘要”工作表，便于在需求评审和质量门禁中直接使用。
"""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Dict, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font


ALLOWED_RESULTS = {"满足", "不满足", "部分满足", "未涉及"}


@dataclass
class ChecklistItem:
    item_id: str
    category: str
    text: str
    row: int


def load_json(path: str | Path) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_items(template_path: str | Path) -> Iterable[ChecklistItem]:
    workbook = load_workbook(template_path)
    worksheet = workbook["安全设计checklist"]

    current_category = ""
    for row in range(1, worksheet.max_row + 1):
        category_cell = worksheet.cell(row, 1).value
        text_cell = worksheet.cell(row, 2).value

        if (
            isinstance(category_cell, str)
            and category_cell.strip()
            and category_cell.strip() not in {"安全设计Checklist", "技术项"}
        ):
            current_category = category_cell.strip()

        if isinstance(text_cell, str):
            match = re.match(r"^(\d+\.\d+)\s+(.*)$", text_cell.strip(), re.S)
            if match:
                yield ChecklistItem(
                    item_id=match.group(1),
                    category=current_category,
                    text=match.group(2).strip(),
                    row=row,
                )


def validate_results(items: Iterable[ChecklistItem], review_results: Dict[str, Dict[str, str]]) -> None:
    item_ids = [item.item_id for item in items]
    missing = [item_id for item_id in item_ids if item_id not in review_results]
    if missing:
        preview = ", ".join(missing[:10])
        raise ValueError(f"缺少评审结果，共 {len(missing)} 项，示例: {preview}")

    for item_id, payload in review_results.items():
        result = str(payload.get("result", "")).strip()
        if item_id in item_ids and result not in ALLOWED_RESULTS:
            raise ValueError(
                f"评审项 {item_id} 的结果值非法: {result}。"
                f"仅允许: {', '.join(sorted(ALLOWED_RESULTS))}"
            )


def build_summary(items: Iterable[ChecklistItem], review_results: Dict[str, Dict[str, str]]):
    summary = {
        "total": 0,
        "满足": 0,
        "不满足": 0,
        "部分满足": 0,
        "未涉及": 0,
        "by_category": {},
    }

    for item in items:
        payload = review_results[item.item_id]
        result = str(payload.get("result", "")).strip()

        summary["total"] += 1
        summary[result] += 1

        category = item.category or "未分类"
        if category not in summary["by_category"]:
            summary["by_category"][category] = {
                "total": 0,
                "满足": 0,
                "不满足": 0,
                "部分满足": 0,
                "未涉及": 0,
            }

        summary["by_category"][category]["total"] += 1
        summary["by_category"][category][result] += 1

    return summary


def write_summary_sheet(workbook, project_info: Dict, summary: Dict) -> None:
    if "评审摘要" in workbook.sheetnames:
        del workbook["评审摘要"]

    worksheet = workbook.create_sheet("评审摘要")
    bold = Font(bold=True)

    worksheet["A1"] = "安全设计评审摘要"
    worksheet["A1"].font = Font(bold=True, size=14)

    info_rows = [
        ("项目名称", project_info.get("project_name", "")),
        ("评审人", project_info.get("reviewer", "")),
        ("评审日期", project_info.get("review_date", datetime.now().strftime("%Y-%m-%d"))),
        ("来源文档", project_info.get("source_docs", "")),
        ("评审结论", project_info.get("conclusion", "")),
    ]

    current_row = 3
    for label, value in info_rows:
        worksheet.cell(current_row, 1).value = label
        worksheet.cell(current_row, 1).font = bold
        worksheet.cell(current_row, 2).value = value
        current_row += 1

    current_row += 1
    worksheet.cell(current_row, 1).value = "总体统计"
    worksheet.cell(current_row, 1).font = bold
    current_row += 1

    headers = ["总项数", "满足", "部分满足", "不满足", "未涉及"]
    for col, header in enumerate(headers, start=1):
        worksheet.cell(current_row, col).value = header
        worksheet.cell(current_row, col).font = bold

    current_row += 1
    worksheet.cell(current_row, 1).value = summary["total"]
    worksheet.cell(current_row, 2).value = summary["满足"]
    worksheet.cell(current_row, 3).value = summary["部分满足"]
    worksheet.cell(current_row, 4).value = summary["不满足"]
    worksheet.cell(current_row, 5).value = summary["未涉及"]

    current_row += 2
    worksheet.cell(current_row, 1).value = "分安全域统计"
    worksheet.cell(current_row, 1).font = bold
    current_row += 1

    category_headers = ["安全域", "总项数", "满足", "部分满足", "不满足", "未涉及"]
    for col, header in enumerate(category_headers, start=1):
        worksheet.cell(current_row, col).value = header
        worksheet.cell(current_row, col).font = bold

    for category, values in summary["by_category"].items():
        current_row += 1
        worksheet.cell(current_row, 1).value = category
        worksheet.cell(current_row, 2).value = values["total"]
        worksheet.cell(current_row, 3).value = values["满足"]
        worksheet.cell(current_row, 4).value = values["部分满足"]
        worksheet.cell(current_row, 5).value = values["不满足"]
        worksheet.cell(current_row, 6).value = values["未涉及"]

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 12
    worksheet.column_dimensions["F"].width = 12


def generate_review_report(
    template_path: str | Path,
    output_path: str | Path,
    review_results: Dict[str, Dict[str, str]],
    project_info: Dict | None = None,
) -> str:
    template_path = Path(template_path)
    output_path = Path(output_path)
    project_info = project_info or {}

    items = list(extract_items(template_path))
    validate_results(items, review_results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["安全设计checklist"]

    for item in items:
        payload = review_results[item.item_id]
        worksheet.cell(item.row, 3).value = str(payload.get("result", "")).strip()
        worksheet.cell(item.row, 4).value = str(payload.get("note", "")).strip()

    summary = build_summary(items, review_results)
    write_summary_sheet(workbook, project_info, summary)

    workbook.save(output_path)
    return str(output_path.resolve())


def main() -> None:
    parser = argparse.ArgumentParser(description="生成安全设计评审 Excel")
    parser.add_argument("--template", required=True, help="Excel 模板路径")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--results-json", required=True, help="评审结果 JSON 路径")
    parser.add_argument("--project-info-json", help="项目信息 JSON 路径")
    args = parser.parse_args()

    review_results = load_json(args.results_json)
    project_info = load_json(args.project_info_json) if args.project_info_json else {}

    output = generate_review_report(
        template_path=args.template,
        output_path=args.output,
        review_results=review_results,
        project_info=project_info,
    )

    output_file = Path(output)
    print(f"✅ Excel 报告已生成: {output}")
    print(f"📦 文件大小: {output_file.stat().st_size} bytes")


if __name__ == "__main__":
    main()
