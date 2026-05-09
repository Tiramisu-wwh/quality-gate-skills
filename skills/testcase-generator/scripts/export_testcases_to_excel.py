#!/usr/bin/env python3
"""Export structured test cases to an Excel workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


DEFAULT_HEADERS = [
    "用例编号",
    "所属模块",
    "用例标题",
    "前置条件",
    "测试数据",
    "测试步骤",
    "预期结果",
    "用例类型",
    "优先级",
    "是否自动化候选",
    "备注",
]

TEAMBITION_TEMPLATE_HEADERS = [
    "标题*",
    "所属分组",
    "维护人",
    "前置条件",
    "步骤描述",
    "预期结果",
    "用例等级",
    "用例类型",
    "自动化是否完成",
]

TEAMBITION_HEADER_ALIASES = {
    "标题*": ["标题*", "标题", "用例标题", "用例名称"],
    "所属分组": ["所属分组", "所属模块", "模块", "所属模块/页面"],
    "维护人": ["维护人", "负责人", "Owner"],
    "前置条件": ["前置条件", "执行前准备", "执行前准备（可选）"],
    "步骤描述": ["步骤描述", "测试步骤", "执行步骤"],
    "预期结果": ["预期结果", "断言要点"],
    "用例等级": ["用例等级", "优先级"],
    "用例类型": ["用例类型"],
    "自动化是否完成": ["自动化是否完成", "是否自动化候选", "是否建议 Midscene 自动化"],
}

NUMBERED_LINE_PATTERN = re.compile(r"^\s*(?:【\d+】|\d+[\.、])")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="JSON payload path")
    parser.add_argument("--output", required=True, help="Excel output path")
    parser.add_argument("--sheet-name", default="测试用例", help="Worksheet name")
    parser.add_argument(
        "--template",
        help="Optional workbook template path. Falls back to a new workbook if invalid.",
    )
    return parser.parse_args()


def load_payload(path: Path) -> tuple[str, list[str], list[list[Any]]]:
    data = json.loads(path.read_text(encoding="utf-8"))

    if isinstance(data, list):
        headers, rows = normalize_rows(data)
        return "测试用例", headers, rows

    if not isinstance(data, dict):
        raise ValueError("Input JSON must be a list or an object.")

    sheet_name = str(data.get("sheet_name") or "测试用例")
    headers = data.get("headers")
    raw_rows = data.get("rows", [])

    if isinstance(raw_rows, list) and raw_rows and isinstance(raw_rows[0], dict):
        inferred_headers, rows = normalize_rows(raw_rows, headers=headers)
        return sheet_name, inferred_headers, rows

    if not headers:
        headers = DEFAULT_HEADERS
    rows = [list(row) for row in raw_rows]
    return sheet_name, list(headers), rows


def normalize_rows(
    rows: list[dict[str, Any]], headers: list[str] | None = None
) -> tuple[list[str], list[list[Any]]]:
    ordered_headers = list(headers or [])
    if not ordered_headers:
        seen: set[str] = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    ordered_headers.append(key)

    normalized = []
    for row in rows:
        normalized.append([row.get(header, "") for header in ordered_headers])
    return ordered_headers or DEFAULT_HEADERS, normalized


def open_workbook(template: str | None) -> Workbook:
    if not template:
        return Workbook()
    try:
        return load_workbook(template)
    except Exception:
        return Workbook()


def is_teambition_template(ws) -> bool:
    values = [ws.cell(1, col).value for col in range(1, len(TEAMBITION_TEMPLATE_HEADERS) + 1)]
    return values == TEAMBITION_TEMPLATE_HEADERS


def rows_to_dicts(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    return [
        {header: row[idx] if idx < len(row) else "" for idx, header in enumerate(headers)}
        for row in rows
    ]


def stringify_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join("" if item is None else str(item) for item in value)
    return value


def find_value(row: dict[str, Any], aliases: list[str]) -> Any:
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return stringify_value(row[alias])
    return ""


def split_non_empty_lines(value: Any) -> list[str]:
    text = stringify_value(value)
    return [line.strip() for line in str(text).splitlines() if line.strip()]


def is_numbered_lines(lines: list[str]) -> bool:
    return bool(lines) and all(NUMBERED_LINE_PATTERN.match(line) for line in lines)


def normalize_numbered_lines(value: Any) -> str:
    lines = split_non_empty_lines(value)
    if not lines:
        return ""
    if len(lines) == 1 or is_numbered_lines(lines):
        return "\n".join(lines)
    return "\n".join(f"【{idx}】{line}" for idx, line in enumerate(lines, start=1))


def validate_teambition_row(row: dict[str, Any], row_number: int) -> None:
    title = find_value(row, TEAMBITION_HEADER_ALIASES["标题*"])
    if str(title).strip() == "":
        raise ValueError(f"第 {row_number} 条用例缺少标题，标题为必填项。")

    steps = split_non_empty_lines(find_value(row, TEAMBITION_HEADER_ALIASES["步骤描述"]))
    expected = split_non_empty_lines(find_value(row, TEAMBITION_HEADER_ALIASES["预期结果"]))
    if len(steps) > 1 or len(expected) > 1:
        if len(steps) != len(expected):
            raise ValueError(
                f"第 {row_number} 条用例的操作步骤与预期结果行数不一致，请按一条步骤对应一条预期结果填写。"
            )


def build_teambition_row(row: dict[str, Any], row_number: int) -> dict[str, Any]:
    validate_teambition_row(row, row_number)
    normalized = dict(row)
    normalized["标题*"] = find_value(row, TEAMBITION_HEADER_ALIASES["标题*"])
    normalized["步骤描述"] = normalize_numbered_lines(
        find_value(row, TEAMBITION_HEADER_ALIASES["步骤描述"])
    )
    normalized["预期结果"] = normalize_numbered_lines(
        find_value(row, TEAMBITION_HEADER_ALIASES["预期结果"])
    )
    return normalized


def autosize_columns(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, max((len(line) for line in value.splitlines()), default=0))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 4, 12), 60)


def write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.delete_rows(1, ws.max_row)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    autosize_columns(ws)


def write_teambition_template_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    row_dicts = rows_to_dicts(headers, rows)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_idx, row in enumerate(row_dicts, start=2):
        normalized_row = build_teambition_row(row, row_idx - 1)
        for col_idx, template_header in enumerate(TEAMBITION_TEMPLATE_HEADERS, start=1):
            target = ws.cell(row=row_idx, column=col_idx)
            target.value = find_value(normalized_row, TEAMBITION_HEADER_ALIASES[template_header])
            target.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    payload_sheet_name, headers, rows = load_payload(input_path)

    workbook = open_workbook(args.template)
    if workbook.sheetnames:
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        worksheet = workbook.active

    if is_teambition_template(worksheet):
        write_teambition_template_sheet(worksheet, headers, rows)
    else:
        worksheet.title = args.sheet_name or payload_sheet_name
        write_sheet(worksheet, headers, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
