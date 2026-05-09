import json
import re
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "export_testcases_to_excel.py"
TEMPLATE = ROOT / "assets" / "测试用例模板.xlsx"


class ExportTestcasesToExcelTests(unittest.TestCase):
    def test_preserves_header_row_and_maps_common_fields(self) -> None:
        payload = {
            "sheet_name": "忽略此值",
            "rows": [
                {
                    "用例标题": "待办列表显示未处理任务",
                    "所属模块": "Web端测试用例|首页|我的待办",
                    "前置条件": "我的待办中存在待处理的任务",
                    "测试步骤": "【1】打开首页\n【2】进入我的待办",
                    "预期结果": "【1】显示未处理任务\n【2】可以进入任务详情",
                    "优先级": "P1",
                    "用例类型": "功能测试",
                    "是否自动化候选": "是",
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_path = tmpdir_path / "payload.json"
            output_path = tmpdir_path / "result.xlsx"
            input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            subprocess.run(
                [
                    "python3",
                    str(SCRIPT),
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                    "--template",
                    str(TEMPLATE),
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            workbook = load_workbook(output_path)
            worksheet = workbook[workbook.sheetnames[0]]

            self.assertEqual(worksheet.title, "template")
            self.assertEqual(worksheet["A1"].value, "标题*")
            self.assertEqual(worksheet["A2"].value, "待办列表显示未处理任务")
            self.assertEqual(worksheet["B2"].value, "Web端测试用例|首页|我的待办")
            self.assertIsNone(worksheet["C2"].value)
            self.assertEqual(worksheet["D2"].value, "我的待办中存在待处理的任务")
            self.assertEqual(worksheet["E2"].value, "【1】打开首页\n【2】进入我的待办")
            self.assertEqual(worksheet["F2"].value, "【1】显示未处理任务\n【2】可以进入任务详情")
            self.assertEqual(worksheet["G2"].value, "P1")
            self.assertEqual(worksheet["H2"].value, "功能测试")
            self.assertEqual(worksheet["I2"].value, "是")

    def test_requires_title_for_teambition_template(self) -> None:
        payload = {
            "rows": [
                {
                    "所属模块": "Web端测试用例|首页|我的待办",
                    "测试步骤": "打开首页",
                    "预期结果": "显示待办列表",
                    "优先级": "P1",
                    "用例类型": "功能测试",
                }
            ]
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_path = tmpdir_path / "payload.json"
            output_path = tmpdir_path / "result.xlsx"
            input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            result = subprocess.run(
                [
                    "python3",
                    str(SCRIPT),
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                    "--template",
                    str(TEMPLATE),
                ],
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(result.returncode, 0)
            self.assertRegex(result.stderr, re.compile("标题.*必填"))

    def test_normalizes_multiline_steps_and_expected_results(self) -> None:
        payload = {
            "rows": [
                {
                    "用例标题": "待办列表支持查看详情",
                    "所属模块": "Web端测试用例|首页|我的待办",
                    "测试步骤": "打开首页\n进入我的待办",
                    "预期结果": "显示待办列表\n可以进入任务详情",
                    "优先级": "P1",
                    "用例类型": "功能测试",
                }
            ]
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_path = tmpdir_path / "payload.json"
            output_path = tmpdir_path / "result.xlsx"
            input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            subprocess.run(
                [
                    "python3",
                    str(SCRIPT),
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                    "--template",
                    str(TEMPLATE),
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            workbook = load_workbook(output_path)
            worksheet = workbook[workbook.sheetnames[0]]

            self.assertEqual(worksheet["E2"].value, "【1】打开首页\n【2】进入我的待办")
            self.assertEqual(worksheet["F2"].value, "【1】显示待办列表\n【2】可以进入任务详情")


if __name__ == "__main__":
    unittest.main()
